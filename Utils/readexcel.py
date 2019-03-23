import openpyxl
def read_data():

    wb=openpyxl.load_workbook('C:/Users/Dusa/PycharmProjects/FreeCRM/Utils/inf.xlsx')
    ws=wb['Sheet1']
    row=ws.max_row
    l=[]

    for i in range(1,row+1):
        l1=[]
        username=ws.cell(i,1)
        password=ws.cell(i,2)
        l1.insert(0,username.value)

        l1.insert(1,password.value)
        l.insert(i-1,l1)
    print(l)
    return l
read_data()